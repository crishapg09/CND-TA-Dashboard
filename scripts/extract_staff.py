#!/usr/bin/env python3
"""
Extract the CND staff roster into the app's data file.

Usage:
    python scripts/extract_staff.py path/to/CND_staff.xlsx

What it does:
  1. Reads the "CND_staff" export (first sheet).
  2. Writes app/src/data/staff.json — one record per staff member (minified).

The roster links to the TA cases on staff Name = case "Assigned to" (the case
`lead`). Names, titles, thematic areas and locations are trimmed of the stray
trailing whitespace the source export carries so the join is exact.

Requires: openpyxl  (pip install openpyxl)
"""
import sys
import os
import json

# ---- source layout -------------------------------------------------------
# Header row followed by one row per staff member.
EXPECTED_HEADER = ('Name', 'Title', 'Thematic Area', 'Location')

C_NAME, C_TITLE, C_AREA, C_LOCATION = 0, 1, 2, 3


def s(v):
    """Trimmed string; None -> ''. Collapses the export's stray trailing spaces."""
    return '' if v is None else str(v).strip()


# ---- duty-station enrichment ------------------------------------------------
# The source HR roster leaves most Location cells blank, which hid where TA is
# actually delivered from. These duty stations were reconstructed from the CND
# team's "Where support flows" map (built in Claude Design): each TA lead's base
# was recovered, resolving the origin for 334 of 349 requests (up from 141).
# Applied on top of the roster's own Location so the join reflects the real
# delivery geography. Names not listed keep their roster Location.
DUTY_STATION = {
    'Amirhossein Yarparvar': 'Amman', 'Annalies Borrel': 'Amman',
    'Karan Courtney Haag': 'Amman', 'Kathleen Heneghan': 'Amman',
    'Mueni Mutunga': 'Amman', 'Odai Abdel Rahman': 'Amman',
    'Alex Mokori': 'Bangkok', 'Christiane Rudert': 'Bangkok',
    'Rene Gerard Galera': 'Bangkok', 'Vani Sethi': 'Bangkok', 'Zivai Murira': 'Bangkok',
    'Federica Margini': 'Brussels', 'Katherine Shats': 'Brussels', 'Roland Kupka': 'Brussels',
    'Aashima Garg': 'Nairobi', 'Agnes Erzse': 'Nairobi', 'Alberto Musatti': 'Nairobi',
    'Alison Feeley': 'Nairobi', 'Ann Defraye': 'Nairobi', 'Anuradha Narayan': 'Nairobi',
    'Benjamin Guy Stafford Allen': 'Nairobi', 'Boniface Kakhobwe': 'Nairobi',
    'Jecinter Akinyi Oketch': 'Nairobi', 'Joan Matji': 'Nairobi', 'Linda Shaker': 'Nairobi',
    'Louise Mwirigi': 'Nairobi', 'Manpreet Kaur Chadha': 'Nairobi', 'Marjorie Volege': 'Nairobi',
    'Mauro Brero': 'Nairobi', 'Minh Tram Le': 'Nairobi', 'Najwa Al Dheeb': 'Nairobi',
    'Nkeiruka Enwelum': 'Nairobi', 'Rowena Katherine Merritt': 'Nairobi',
    'Simeon Nanama': 'Nairobi', 'Zephenia Gomora': 'Nairobi',
    'Nita Dalmiya': 'New York', 'Amal Ben Ameur': 'Panama', 'Paula Veliz': 'Panama',
}


def load_rows(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    ws.reset_dimensions()
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    header = tuple(s(v) for v in rows[0][:len(EXPECTED_HEADER)])
    if header != EXPECTED_HEADER:
        raise SystemExit(
            'ERROR: unexpected staff layout.\n'
            f'  expected {EXPECTED_HEADER}\n'
            f'  got      {header}\n'
            'The column mapping in this script would need updating.'
        )
    return rows[1:]


def build(r):
    """One roster row -> one Staff record (see app/src/data/types.ts)."""
    name = s(r[C_NAME])
    location = s(r[C_LOCATION]) if len(r) > C_LOCATION else ''
    return {
        'name': name,
        'title': s(r[C_TITLE]) if len(r) > C_TITLE else '',
        'area': s(r[C_AREA]) if len(r) > C_AREA else '',
        'location': DUTY_STATION.get(name, location),
    }


def main():
    if len(sys.argv) != 2:
        raise SystemExit('Usage: python scripts/extract_staff.py <CND_staff.xlsx>')
    src = sys.argv[1]
    repo_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    data_dir = os.path.join(repo_root, 'app', 'src', 'data')

    rows = load_rows(src)
    records = [build(r) for r in rows if r and s(r[C_NAME])]
    # De-duplicate on name, keeping the first occurrence.
    seen = set()
    unique = []
    for rec in records:
        if rec['name'] in seen:
            continue
        seen.add(rec['name'])
        unique.append(rec)

    with open(os.path.join(data_dir, 'staff.json'), 'w', encoding='utf-8') as f:
        json.dump(unique, f, ensure_ascii=False, separators=(',', ':'))

    from collections import Counter
    areas = dict(Counter(r['area'] for r in unique))
    print(f'Wrote {len(unique):,} staff records')
    print(f'Thematic areas: {areas}')


if __name__ == '__main__':
    main()
